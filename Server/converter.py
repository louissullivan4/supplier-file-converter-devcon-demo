import openpyxl
from openpyxl import load_workbook
import string


def create_columns(output_sheet):
    column_list = [
        "sku", "product_type", "categories", "product_websites", "name", "description", "colour", "weight",
        "product_online",
        "tax_class_name", "visibility", "price", "qty"]
    for index in range(len(column_list)):
        cell = output_sheet.cell(row=1, column=(index + 1))
        cell.value = column_list[index]


def fill_known(input_sheet, output_sheet):
    # dict layout is name of column as key and then the two values are input_col and output_col
    cols_dict = {"sku": [1, 1], "categories": [12, 3], "name": [2, 5], "description": [11, 6],
                 "colour": [5, 7], "size": [8, 8], "price": [18, 12], "quantity": [19, 13]}
    product_count = 0
    for key, val in cols_dict.items():
        for row in range(2, input_sheet.max_row):
            if input_sheet.cell(row, 1).value is None:
                product_count = row
                break
            cell = output_sheet.cell(row=row, column=val[1])
            if key == "name":
                cell_val = input_sheet.cell(row, val[0]).value
                cell.value = string.capwords(cell_val)
            else:
                cell.value = input_sheet.cell(row, val[0]).value
    return product_count


def fill_others(product_count, output_sheet):
    product_type = 2
    product_websites = 4
    weight = 8
    product_online = 9
    tax_class_name = 10
    not_visible = 11
    for row in range(2, product_count):
        cell = output_sheet.cell(row=row, column=product_type)
        cell.value = "simple"
        cell = output_sheet.cell(row=row, column=product_websites)
        cell.value = "base"
        cell = output_sheet.cell(row=row, column=weight)
        cell.value = "1"
        cell = output_sheet.cell(row=row, column=product_online)
        cell.value = "1"
        cell = output_sheet.cell(row=row, column=tax_class_name)
        cell.value = "Taxable Goods"
        cell = output_sheet.cell(row=row, column=not_visible)
        cell.value = "Not Visible Individually"


def convert(filelocation):
    input_workbook = load_workbook(filename=filelocation)
    output_workbook = openpyxl.Workbook()
    input_sheet = input_workbook.active
    output_sheet = output_workbook.active
    create_columns(output_sheet)
    product_count = fill_known(input_sheet, output_sheet)
    fill_others(product_count, output_sheet)
    output_workbook.save(filename="static/output/output.xlsx")


if __name__ == '__main__':
    convert("static/uploads/input.xlsx")
